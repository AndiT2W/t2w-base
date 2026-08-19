import csv, json, re
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

src=Path(r'C:\Users\andi\Downloads\2026-08-19T14_18_42.962Z TIME 2 WIN - TIME 2 WIN - Office - VERANSTALTUNGEN.xlsx')
out=Path(r'C:\work\t2w-base\outputs\clickup-import'); out.mkdir(parents=True,exist_ok=True)
wb=openpyxl.load_workbook(src,data_only=True,read_only=True)
ws=wb['Tasks']; it=ws.iter_rows(values_only=True)
for _ in range(3): next(it)
headers=list(next(ws.iter_rows(min_row=3,max_row=3,values_only=True)))
headers=[str(h or '').strip() for h in headers]
rows=[]
for vals in it:
    if not any(v not in (None,'') for v in vals): continue
    row=dict(zip(headers, vals))
    if not row.get('Task ID') or not row.get('Task Name'): continue
    rows.append(row)

def snake(s):
    s=s.replace('�','ae').replace('ä','ae').replace('ö','oe').replace('ü','ue').replace('Ä','Ae').replace('Ö','Oe').replace('Ü','Ue').replace('ß','ss')
    s=re.sub(r"[^A-Za-z0-9]+", '_', s).strip('_').lower()
    return s or 'field'
keys=[]; seen=set()
for h in headers:
    k=snake(h)
    if k in seen: i=2; base=k
    else: i=1; base=k
    while k in seen: k=f'{base}_{i}'; i+=1
    seen.add(k); keys.append(k)

def norm(v):
    if isinstance(v, datetime): return v.isoformat()
    return v
norm_rows=[[norm(r.get(h,'')) for h in headers] for r in rows]
# CSV
csv_path=out/'t2w-events-import.csv'
with csv_path.open('w',newline='',encoding='utf-8-sig') as f:
    w=csv.writer(f); w.writerow(keys); w.writerows(norm_rows)
# XLSX
x=Workbook(); sh=x.active; sh.title='events'; sh.append(keys)
for row in norm_rows: sh.append(row)
header_fill=PatternFill('solid',fgColor='1F4E78')
for c in sh[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=header_fill; c.alignment=Alignment(wrap_text=True)
sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
for col in sh.columns:
    letter=col[0].column_letter; sh.column_dimensions[letter].width=min(max(max(len(str(c.value or '')) for c in col[:30])+2,12),38)
# mapping
mp=x.create_sheet('field_mapping'); mp.append(['source_header','import_header','custom_field'])
for h,k in zip(headers,keys): mp.append([h,k, h not in {'Task Type','Task ID','Task Name','Parent ID','Parent Name','Parent URL','Status','Task Content','Assignee','Priority','Latest Comment','Comment Count','Assigned Comment Count','Due Date','Start Date','Date Created','Date Updated','Date Closed','Date Done','Created By','Space','Folder','List'}])
for c in mp[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=header_fill
mp.freeze_panes='A2'; mp.auto_filter.ref=mp.dimensions
for col in mp.columns:
    mp.column_dimensions[col[0].column_letter].width=42
x.save(out/'t2w-events-import.xlsx')
print(json.dumps({'rows':len(rows),'columns':len(headers),'csv':str(csv_path),'xlsx':str(out/'t2w-events-import.xlsx')}))
